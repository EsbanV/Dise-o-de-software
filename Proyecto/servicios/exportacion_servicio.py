from modelos import Usuario, CuentaBancaria, Categoria, Transaccion
import pandas as pd
import io
from openpyxl.styles import Font, Alignment

class ExportacionServicio:
    def __init__(self, transaccion_repositorio):
        self.transaccion_repositorio = transaccion_repositorio

    def exportar_excel(self, usuario_id, cuenta_id):
        usuario = self.transaccion_repositorio.obtener_por_id(Usuario, usuario_id)
        if not usuario:
            raise LookupError("Usuario no encontrado")

        cuenta = self.transaccion_repositorio.obtener_unico_con_filtro(
            CuentaBancaria, [
                CuentaBancaria.id == cuenta_id,
                CuentaBancaria.usuario_id == usuario_id
            ]
        )
        if not cuenta:
            raise LookupError("Cuenta no encontrada")

        transacciones = self.transaccion_repositorio.obtener_transacciones_con_categoria(cuenta_id)

        data = [
            {
                "categoria": t.categoria.nombre,
                "tipo": "Ingreso" if t.categoria.tipo.name == "INGRESO" else "Gasto",
                "descripcion": t.descripcion,
                "monto": t.monto,
                "fecha": t.fecha
            }
            for t in transacciones
        ]
        df = pd.DataFrame(data, columns=["categoria", "tipo", "descripcion", "monto", "fecha"])

        total_ingresos = df[df['monto'] > 0]['monto'].sum() if not df.empty else 0
        total_gastos = abs(df[df['monto'] < 0]['monto'].sum()) if not df.empty else 0
        saldo_cuenta = cuenta.saldo
        usuario_nombre = usuario.nombre
        nombre_cuenta = cuenta.nombre

        df['Total'] = df['monto'].cumsum() + (saldo_cuenta - df['monto'].sum() if not df.empty else saldo_cuenta)
        df = df[['categoria', 'tipo', 'descripcion', 'monto', 'fecha', 'Total']]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Transacciones')
            ws = writer.sheets['Transacciones']

            # Formato columnas
            column_widths = {
                'A': 20, 'B': 15, 'C': 30, 'D': 15, 'E': 20, 'F': 15
            }
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            for col in ['D', 'F']:
                for row in range(2, len(df) + 2):
                    cell = ws[f'{col}{row}']
                    cell.number_format = '"₡"#,##0.00'
            for row in range(2, len(df) + 2):
                monto_cell = ws[f'D{row}']
                if monto_cell.value and monto_cell.value < 0:
                    monto_cell.font = Font(color="FF0000")
            ws.insert_rows(1, 4)
            ws['A1'] = "Usuario:"; ws['B1'] = usuario_nombre
            ws['A2'] = "Cuenta:"; ws['B2'] = nombre_cuenta
            ws['A3'] = "Saldo Actual:"; ws['B3'] = saldo_cuenta
            ws['B3'].number_format = '"₡"#,##0.00'
            ws['A4'] = "Total Ingresos:"; ws['B4'] = total_ingresos
            ws['B4'].number_format = '"₡"#,##0.00'
            ws['D4'] = "Total Gastos:"; ws['E4'] = total_gastos
            ws['E4'].number_format = '"₡"#,##0.00'
            ws['E4'].font = Font(color="FF0000")
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            ws.freeze_panes = 'A6'

        output.seek(0)
        return output, usuario_nombre, nombre_cuenta
