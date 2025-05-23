from flask import Blueprint, send_file, session, request, redirect, url_for
import sqlite3
import pandas as pd
import io
from pathlib import Path
from datetime import datetime
from openpyxl.styles import Font, Alignment, numbers
from openpyxl.utils import get_column_letter

exportacion_rutas = Blueprint('exportacion_rutas', __name__)

@exportacion_rutas.route('/exportar_excel')
def exportar_excel():
    try:
        usuario_id = session.get('usuario_id')
        if not usuario_id:
            return redirect(url_for('usuario_rutas.login'))

        cuenta_id = request.args.get('cuenta_id', type=int)
        if not cuenta_id:
            return "Se requiere el ID de la cuenta", 400

        db_path = Path(__file__).parent.parent / 'instance' / 'base_datos.db'
        
        conn = sqlite3.connect(db_path)
        
        # Obtener información del usuario y cuenta
        usuario_query = "SELECT nombre FROM usuarios WHERE id = ?"
        usuario_nombre = conn.execute(usuario_query, (usuario_id,)).fetchone()[0]
        
        cuenta_query = "SELECT nombre, saldo FROM cuentas_bancarias WHERE id = ? AND usuario_id = ?"
        cuenta_data = conn.execute(cuenta_query, (cuenta_id, usuario_id)).fetchone()
        
        if not cuenta_data:
            conn.close()
            return "Cuenta no encontrada", 404

        nombre_cuenta, saldo_cuenta = cuenta_data
        
        # Obtener transacciones
        transacciones_query = """
        SELECT 
            c.nombre as categoria,
            c.tipo,
            t.descripcion,
            t.monto,
            t.fecha
        FROM transacciones t
        JOIN categorias c ON t.categoria_id = c.id
        WHERE t.cuenta_bancaria_id = ?
        ORDER BY t.fecha DESC
        """
        df = pd.read_sql_query(transacciones_query, conn, params=(cuenta_id,))
        
        # Calcular totales
        total_ingresos = df[df['monto'] > 0]['monto'].sum()
        total_gastos = df[df['monto'] < 0]['monto'].sum() * -1
        
        # Crear DataFrame con columna de Total Acumulado
        df['Total'] = df['monto'].cumsum() + (saldo_cuenta - df['monto'].sum())
        
        # Reordenar columnas para poner Total junto a Fecha
        df = df[['categoria', 'tipo', 'descripcion', 'monto', 'fecha', 'Total']]
        
        conn.close()

        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja de transacciones
            df.to_excel(writer, index=False, sheet_name='Transacciones')
            
            # Obtener el libro y hoja para formato
            workbook = writer.book
            ws = writer.sheets['Transacciones']
            
            # Configurar anchos de columnas
            column_widths = {
                'A': 20,  # Categoría
                'B': 15,  # Tipo
                'C': 30,  # Descripción
                'D': 15,  # Monto
                'E': 20,  # Fecha
                'F': 15   # Total
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # Formato de moneda para columnas D (Monto) y F (Total)
            for col in ['D', 'F']:
                for row in range(2, len(df)+2):  # +2 porque empieza en fila 2 (1 es header)
                    cell = ws[f'{col}{row}']
                    cell.number_format = '"₡"#,##0.00'
            
            # Formato para montos positivos/negativos
            for row in range(2, len(df)+2):
                monto_cell = ws[f'D{row}']
                if monto_cell.value and monto_cell.value < 0:
                    monto_cell.font = Font(color="FF0000")  # Rojo para gastos
            
            # Añadir información de cabecera
            ws.insert_rows(1, 4)  # Insertar 4 filas al inicio
            
            # Información del usuario y cuenta
            ws['A1'] = "Usuario:"
            ws['B1'] = usuario_nombre
            ws['A2'] = "Cuenta:"
            ws['B2'] = nombre_cuenta
            ws['A3'] = "Saldo Actual:"
            ws['B3'] = saldo_cuenta
            ws['B3'].number_format = '"₡"#,##0.00'
            
            # Totales
            ws['A4'] = "Total Ingresos:"
            ws['B4'] = total_ingresos
            ws['B4'].number_format = '"₡"#,##0.00'
            ws['D4'] = "Total Gastos:"
            ws['E4'] = total_gastos
            ws['E4'].number_format = '"₡"#,##0.00'
            ws['E4'].font = Font(color="FF0000")
            
            # Estilo para la cabecera
            for cell in ws[5]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Congelar paneles para ver cabeceras al desplazar
            ws.freeze_panes = 'A6'

        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f'reporte_{usuario_nombre}_{nombre_cuenta}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except sqlite3.Error as e:
        print(f"Error de base de datos: {str(e)}")
        return f"Error al acceder a la base de datos: {str(e)}", 500
    except Exception as e:
        print(f"Error inesperado: {str(e)}")
        return f"Error interno al generar el reporte: {str(e)}", 500